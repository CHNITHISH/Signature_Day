from flask import Flask, render_template, request, redirect, url_for
import razorpay
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

razorpay_client = razorpay.Client(auth=("rzp_test_tBShEJdYWLgAzr", "ymsNG0fQ9yE9onZ88JFHFGZ9"))

if not os.path.exists('data.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.append(['Name', 'Roll Number', 'Size', 'Image Path'])
    wb.save('data.xlsx')

@app.route('/')
def index():
    return render_template('form.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    roll = int(request.form['roll'])
    size = request.form['size']

    if roll > 78 or size not in ['XS', 'S', 'M', 'L', 'XL', 'XXL']:
        return 'Invalid roll number or size!'

    image = request.files['image']
    ext = os.path.splitext(image.filename)[1]
    image_filename = f"{roll}{ext}"
    image_path = os.path.join(app.config['UPLOAD_FOLDER'], image_filename)
    image.save(image_path)

    wb = load_workbook('data.xlsx')
    ws = wb.active
    ws.append([name, roll, size, image_path])
    data = sorted(ws.iter_rows(min_row=2, values_only=True), key=lambda x: x[1])
    ws.delete_rows(2, ws.max_row)
    for row in data:
        ws.append(row)
    wb.save('data.xlsx')

    return redirect(url_for('success'))

@app.route('/success')
def success():
    return render_template('success.html')

@app.route('/create_order', methods=['POST'])
def create_order():
    amount = 150 * 100
    payment = razorpay_client.order.create({'amount': amount, 'currency': 'INR', 'payment_capture': '1'})
    return {'order_id': payment['id']}

if __name__ == "__main__":
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run()

